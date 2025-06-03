import csv
import logging
import os
import openpyxl # Mantido para a função de Excel
from openpyxl import utils as openpyxl_utils # Mantido
from datetime import datetime # Mantido

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - (report_generator) - %(message)s')

# --- Suas funções auxiliares (_formatar_competencia_aaaamm, etc.) permanecem aqui como estavam ---
def _formatar_competencia_aaaamm(competencia_aamm_str, data_emissao_yyyymmdd_str):
    if competencia_aamm_str and len(competencia_aamm_str) == 4 and \
       data_emissao_yyyymmdd_str and len(data_emissao_yyyymmdd_str) == 8:
        try:
            ano_emissao_completo = data_emissao_yyyymmdd_str[0:4]
            seculo_emissao = ano_emissao_completo[0:2]
            ano_competencia_curto = competencia_aamm_str[0:2]
            mes_competencia = competencia_aamm_str[2:4]
            return f"{seculo_emissao}{ano_competencia_curto}{mes_competencia}"
        except Exception:
            return competencia_aamm_str
    elif competencia_aamm_str and len(competencia_aamm_str) == 6:
        return competencia_aamm_str
    return competencia_aamm_str

def _formatar_data_para_relatorio(data_str_yyyymmdd):
    if not data_str_yyyymmdd or len(data_str_yyyymmdd) != 8:
        return data_str_yyyymmdd
    try:
        dt_obj = datetime.strptime(data_str_yyyymmdd, '%Y%m%d')
        return dt_obj.strftime('%d/%m/%Y')
    except ValueError:
        return data_str_yyyymmdd

def _formatar_valor_para_numero(valor_str):
    if valor_str is None:
        return 0.0
    try:
        valor_corrigido_str = str(valor_str).replace(',', '.')
        return float(valor_corrigido_str)
    except ValueError:
        logging.warning(f"Não foi possível converter valor '{valor_str}' para número. Usando 0.0.")
        return 0.0

def gerar_relatorio_distribuicao(plano_distribuicao, caminho_pasta_distribuicao):
    # ... (seu código para gerar_relatorio_distribuicao permanece o mesmo que você me passou) ...
    # Apenas garanta que os logging.info, logging.error etc. estejam sendo usados
    # em vez de print(), como já ajustamos.
    if not plano_distribuicao:
        logging.error("Plano de distribuição está vazio. Relatório Excel não gerado.")
        return False, None

    nome_arquivo_excel = "DISTRIBUIÇÃO.xlsx"
    caminho_completo_excel = os.path.join(caminho_pasta_distribuicao, nome_arquivo_excel)

    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active 

        if sheet is None: 
            logging.error("Não foi possível obter a planilha ativa do novo workbook Excel.")
            return False, None
        
        sheet.title = "Distribuição Faturas Audit+" 

        cabecalhos = [
            "Nº FATURA", "COMP", "UNIMED", 
            "EMISSÃO", "VENCIMENTO", "VALOR", 
            "AUDITOR"
        ]
        sheet.append(cabecalhos)

        for nome_auditor, dados_auditor in plano_distribuicao.items():
            for fatura_info in dados_auditor.get('faturas', []):
                num_fatura = fatura_info.get('numero_fatura', 'N/A')
                
                competencia_original = fatura_info.get('competencia', 'N/A')
                data_emissao_original = fatura_info.get('data_emissao', '')
                competencia_fmt = _formatar_competencia_aaaamm(competencia_original, data_emissao_original)

                cod_uni_destino = fatura_info.get('codigo_unimed_destino', '')
                nome_uni_destino = fatura_info.get('nome_unimed_destino', 'N/A')
                
                if cod_uni_destino and nome_uni_destino and "não encontrada" not in nome_uni_destino.lower() and nome_uni_destino != 'N/A':
                    unimed_destino_formatada = f"{cod_uni_destino} - {nome_uni_destino}"
                elif cod_uni_destino:
                    unimed_destino_formatada = f"{cod_uni_destino} - (Nome não localizado)"
                else:
                    unimed_destino_formatada = nome_uni_destino

                data_emissao_fmt = _formatar_data_para_relatorio(data_emissao_original)
                data_vencimento_fmt = _formatar_data_para_relatorio(fatura_info.get('data_vencimento', ''))
                valor_liquido_num = _formatar_valor_para_numero(fatura_info.get('valor_total_documento', '0'))

                linha_dados = [
                    num_fatura, competencia_fmt, unimed_destino_formatada,
                    data_emissao_fmt, data_vencimento_fmt, valor_liquido_num,
                    nome_auditor
                ]
                sheet.append(linha_dados)

                if isinstance(valor_liquido_num, float):
                    cell = sheet.cell(row=sheet.max_row, column=cabecalhos.index("VALOR") + 1)
                    cell.number_format = 'R$ #,##0.00' 
        
        for col_idx, column_cells in enumerate(sheet.columns): 
            if sheet is None: break 
            max_length = 0
            column_letter = openpyxl_utils.get_column_letter(col_idx + 1) 
            for cell in column_cells:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length > 0 else 12 
            if sheet.column_dimensions is not None: 
                    sheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(filename=caminho_completo_excel)
        logging.info(f"Relatório '{nome_arquivo_excel}' gerado com sucesso em '{caminho_pasta_distribuicao}'.")
        return True, caminho_completo_excel

    except Exception as e:
        logging.exception(f"Falha ao gerar o relatório Excel. Erro: {e}")
        return False, None


def gerar_csv_internacao(guias_relevantes, output_folder):
    """
    Gera um arquivo CSV com as guias de internação consideradas relevantes.
    """
    if not guias_relevantes:
        logging.warning("Nenhuma guia de internação relevante foi fornecida para o CSV. O arquivo não será gerado.")
        return True

    output_filename = "Guias de Internação Relevantes.csv"
    output_path = os.path.join(output_folder, output_filename)

    headers = [
        "Fatura Pai",
        "Nº Guia Internação",
        "Código Beneficiário",
        "Nome Beneficiário",
        "Tipo de Internação", # ADICIONADO
        "Valor p/ Filtro (R$)",
        "Valor Real Total (R$)"
    ]

    logging.info(f"Gerando CSV de guias de internação em: {output_path}")

    try:
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile, delimiter=';')
            writer.writerow(headers)

            for guia in guias_relevantes:
                valor_filtro_str = "{:.2f}".format(guia.get('valor_filtro', 0.0)).replace('.', ',')
                valor_real_str = "{:.2f}".format(guia.get('valor_total_real', 0.0)).replace('.', ',')

                writer.writerow([
                    guia.get('fatura_pai', ''),
                    guia.get('numero_guia', ''),
                    guia.get('codigo_beneficiario', ''),
                    guia.get('nome_beneficiario', ''),
                    guia.get('tipo_internacao', ''), # ADICIONADO
                    valor_filtro_str,
                    valor_real_str
                ])

        logging.info(f"Arquivo CSV '{output_filename}' gerado com sucesso com {len(guias_relevantes)} guias.")
        return True
    except IOError as e:
        logging.exception(f"Erro de E/S ao tentar escrever o arquivo CSV em {output_path}: {e}")
        return False
    except Exception as e:
        logging.exception(f"Erro inesperado ao gerar o arquivo CSV: {e}")
        return False

if __name__ == '__main__':
    # Mantenha seu bloco de teste como estava, ou adapte para testar ambas as funções
    # ... (seu código de teste if __name__ == '__main__' que você já tinha) ...
    logging.info("Executando report_generator.py como script principal para teste.")
    # (O bloco de teste if __name__ == '__main__' que você me mostrou anteriormente pode ser mantido aqui)
    # Apenas garanta que ele teste ambas as funções se desejar.
    pass